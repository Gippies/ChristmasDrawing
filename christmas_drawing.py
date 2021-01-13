from colorama import init as colorama_init, Fore
from openpyxl import load_workbook, Workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename, asksaveasfilename
import random


def match_people():
    Tk().withdraw()
    fn = askopenfilename(title="Select an Excel file", filetypes=(("Excel Files", "*.xlsx"), ("All Files", "*.*")))

    if fn == '':
        print(Fore.RED + "No file selected. Exiting...")
        return

    wb = load_workbook(fn, read_only=True)
    sheet = wb.worksheets[0]
    print("Calculating worksheet dimensions...")
    sheet_dimensions = sheet.calculate_dimension().split(':')
    if sheet_dimensions[0] != 'A1' or sheet_dimensions[1] == 'A1':
        print(Fore.YELLOW + "Warning: worksheet dimensions appear to be incorrect,"
                            "resetting dimensions and recalculating...")
        sheet.reset_dimensions()
        try:
            sheet.calculate_dimension(force=True)
        except UnboundLocalError:
            print(Fore.RED + "Error: It appears you may be using an empty (or very broken) Excel spreadsheet."
                  "Try adding some data or using one that's not broken. Exiting...")
            return

    print(Fore.GREEN + "Dimensions calculated successfully. Beginning pairing process...")
    exclusion_dict = {}
    for row in sheet.rows:
        if row[0].value in exclusion_dict.keys():
            print(Fore.RED + "Error: There are duplicate individuals in the 'A' column. "
                             "If you have multiple exclusions, please put them all on the same row. Exiting...")
            return
        exclusion_dict[row[0].value] = []
        for cell in row[1:]:
            if cell.value is not None:
                exclusion_dict[row[0].value].append(cell.value)
    wb.close()

    individuals = list(exclusion_dict.keys())
    print("People to pair: " + str(individuals))
    print("Exclusions: " + str(exclusion_dict))

    is_valid = False
    pairs_dict = {}

    while not is_valid:
        is_valid = True
        pairs_dict = {}
        individuals_to_randomize = individuals.copy()
        random.shuffle(individuals_to_randomize)

        for i in range(len(individuals)):
            if individuals[i] == individuals_to_randomize[i] or individuals_to_randomize[i] in exclusion_dict[individuals[i]]:
                is_valid = False
            pairs_dict[individuals[i]] = individuals_to_randomize[i]

        if not is_valid:
            print(Fore.YELLOW + "Warning: Invalid pairs selected. Trying again...")

    print("Shuffled Pairs: " + str(pairs_dict))
    print("Writing to file...")

    fn_to_save = asksaveasfilename(
        filetypes=[('Excel Files', '*.xlsx'), ('All Files', '*.*')],
        defaultextension=[('Excel Files', '*.xlsx'), ('All Files', '*.*')]
    )

    if fn_to_save == '':
        print(Fore.RED + "No save file specified. Exiting...")
        return

    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    for k, v in pairs_dict.items():
        ws.append([k, v])
    wb.save(fn_to_save)


if __name__ == '__main__':
    colorama_init(autoreset=True)
    match_people()
    input("Press Enter to continue...")
